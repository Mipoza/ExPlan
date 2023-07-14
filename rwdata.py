import pandas as pd
import os
import locale
import re
from algo_half import *
from datetime import datetime, timedelta
import ezodf
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

month_fr = {
    "January": "janvier",
    "February": "février",
    "March": "mars",
    "April": "avril",
    "May": "mai",
    "June": "juin",
    "July": "juillet",
    "August": "août",
    "September": "septembre",
    "October": "octobre",
    "November": "novembre",
    "December": "décembre"
}

locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')



def date_index(col, day, date):
    ncol = [str(x).lower() for x in col]
    k = ncol.index(day) #hope there is no issue with "reptitions"
    while not re.match(r"\d{4}-\d{2}-\d{2}|janvier|fevrier|mars|avril|mai|juin|juillet|aout|septembre|octobre|novembre|decembre", str(ncol[k]), re.IGNORECASE):
        k-=1
    return k

def filter(input_string):
    pattern = r"REF[ ]+\d+ ([a-zA-Z ]+) ([\d\+?]+) ([a-zA-Z \d]+) HEB ([a-zA-Z ]*)"

    tab = input_string.split("REF")[1:]  
    tab = ["REF " + e for e in tab]

    results = []

    for s in tab:
        match = re.findall(pattern, s)
        name = match[0][0]
        student_count = match[0][1]
        date_part = match[0][2] #Do i add this to group ?
        heb = match[0][3].rstrip()
        results.append(Group(name,heb,student_count.strip()))


    return results

def check_day(chaine):
    jours_semaine = ["lundi", "mardi", "mercredi", "jeudi", "vendredi", "samedi", "dimanche"]
    pattern = re.compile(r'^(' + '|'.join(jours_semaine) + ')', re.IGNORECASE)
    match = re.match(pattern, chaine)
    if match:
        return match.group(0).capitalize()
    else:
        return None

def k_offset(k,raws):
    week_days = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
    k_0 = k
    day = check_day(str(raws[k_0]))
    if day == None:
        day = check_day(str(raws[k_0-1]))
        if day == None:
            return k_0 #default behavior assuming it will be monday
    return k_0 - 2*week_days.index(day)


    

def get_planning(plan_path, date, bs_list, sp_drivers):
    special_drivers = [d[0] for d in sp_drivers]
    b_list = [b[0] for b in bs_list]
    d_bus = {}
    for e in bs_list:
        d_bus[e[0]] = e[1]

    date_year = date.year
    month_names = [
         "FEVRIER","MARS","AVRIL", "MAI", "JUIN",
        "JUILLET", "AOUT", "SEPTEMBRE"
    ]
    
    sheet_names = [f"{month} {date_year}" for month in month_names]
    data = pd.read_excel(plan_path, sheet_name=sheet_names)

    df = data[sheet_names[date.month-2]] #depend on month_names size
    wb = load_workbook(plan_path)
    sheet = wb[sheet_names[date.month-2]] #to check merged cells
    
    day =  date.strftime('%A') + ' ' + date.strftime('%d') 
    cols = df.columns.tolist()

    k = date_index(df[cols[0]], day.lower(), date)
    staff_name = []
    for c in cols[1:]:
        if df[c][k] == df[c][k]: #check nan values
            staff_name.append(df[c][k])

    cols_s = [c for c in cols if df[c][k] == df[c][k]][1:]

    staff_dict = {key: value for key, value in zip(cols_s, staff_name)}

    input_string = df[cols[1]][k+1]
    groups = filter(input_string)

    complete_plan = []
    buses = b_list.copy()
    check_full = False
    k += 2
    parity = 0
    k_0 = k_offset(k,df[cols[0]])
    while k < len(df[cols[0]]) and not re.match(r"\d{4}-\d{2}-\d{2}|janvier|fevrier|mars|avril|mai|juin|juillet|aout|septembre|octobre|novembre|decembre", str(df[cols[0]][k]), re.IGNORECASE):
        half = df.loc[k, cols_s].tolist()
        works = [c for c in cols_s if df[c][k] == df[c][k] and len(str(df[c][k]).split(',')) >= 3]
        s_list = []
        
        if works: 
            d_day = df[cols[0]][k]
            p_digit = r"\d{2}$"
            d_p = "m" if parity % 2 == 0 else "a"

            if df[cols[0]][k] != df[cols[0]][k]:
                d_day = df[cols[0]][k-1]
                if d_day != d_day:
                    d_day = df[cols[0]][k-2]
            d_day = re.search(p_digit, d_day).group()
            time = d_day + "/" + str(date.month).zfill(2) + "/" + str(date.year) + "/" + d_p
           
            for w in works:
                cell_data = df[w][k].split(',')
                cell_data = [cell.lower().strip() for cell in cell_data]

                s_name = staff_dict[w]
                s_group = [g for g in groups if g.name.lower().strip() == cell_data[0].lower().strip()]
                s_activity = Activity(cell_data[1], cell_data[2], activity_size.get(cell_data[1], 10)) #10 per default value
                s_list.append(Staff(s_name, s_group[0], s_activity, (sheet[k+2][cols.index(w)].coordinate in sheet.merged_cells) and d_p == "m")) #understand/fix th 2 offset

            buses = [b for b in buses if not d_bus[b][k-k_0]]
            b_list_tmp = [b for b in b_list if not d_bus[b][k-k_0]]
            old_size = len(buses)
            plan, buses = planning_halfday(s_list, buses if check_full else b_list_tmp, [special_drivers[i] for i in range(len(special_drivers)) if not sp_drivers[i][1][k-k_0] ], time)
            check_full =  (len(buses) < old_size) and d_p == "m"
            if not check_full:
                buses = b_list.copy()
            complete_plan.append(plan) 
            
        k += 1
        parity += 1

    return complete_plan
    
def write_planning(complete_plan, path):
    # Créer un nouveau classeur Excel
    classeur = Workbook()

    # Sélectionner la sheet active (par défaut, la première sheet)
    sheet = classeur.active

    cols = ['A','B','C','D','E','F','G','H','I','J']

    # Écrire des données dans les cellules
    sheet['A1'] = 'Dates'
    sheet['B1'] = 'Activités'
    sheet['C1'] = 'Lieux de prise en charge'
    sheet['D1'] = 'Groupe'
    sheet['E1'] = 'Moniteurs'
    sheet['F1'] = 'Horaires'
    sheet['G1'] = 'Lieux de dépose'
    sheet['H1'] = 'Véhicules'
    sheet['I1'] = 'Chauffeurs'
    sheet['J1'] = 'Effectif'

    r_c = 2
    for i in range(0,len(complete_plan)):
        
        color = "AAAAAA" if (i > 0 and complete_plan[i-1][0].time[:2] == complete_plan[i][0].time[:2]) else "444444"
        sheet.merge_cells('A'+str(r_c)+':'+cols[-1]+str(r_c))
        sheet['A'+str(r_c)].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        r_c += 1
        p_l = complete_plan[i]
        time = p_l[0].time
        count = len(p_l)
        ma = "matin" if time[-1] == "m" else "après-midi"
        date = datetime.strptime(time[:-2], "%d/%m/%Y")
        sheet['A'+str(r_c)] = date.strftime("%A %d %B") + "\n" + ma
        if len(p_l) > 1:
            sheet.merge_cells('A'+str(r_c)+':'+'A'+str(r_c+len(p_l)-1))

        for j in range(len(p_l)):
            p = p_l[j]
            
            check_full = any(s.full for s in p.staffs)

            sheet['B'+str(r_c+j)] = p.activity.name + (" (journée)" if check_full else "")
            sheet['C'+str(r_c+j)] = p.group.location
            sheet['D'+str(r_c+j)] = p.group.name
            sheet['E'+str(r_c+j)] = '+'.join([s.name for s in p.staffs])
            sheet['F'+str(r_c+j)] = p.schedule
            sheet['G'+str(r_c+j)] = p.activity.location
            sheet['H'+str(r_c+j)] = '+'.join([b.name + " (" + str(b.size) + ")" for b in p.buses])
            sheet['I'+str(r_c+j)] = '+'.join(p.drivers)
            sheet['J'+str(r_c+j)] = p.s_size if p.s_size != "" else p.size
  
        r_c += len(p_l)
    
    """
    for ligne in sheet.iter_rows():
        for cellule in ligne:
            cellule.font = Font(size=14)
    """
    #auto size and align cells

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for c in cols:
        sheet.column_dimensions[c].auto_size = True

    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 0.9
        sheet.column_dimensions[column].width = adjusted_width

    for row in sheet.rows:
        max_height = 0
        for cell in row:
            try:
                cell_height = 25 #make it dynamic and not static
                if cell_height > max_height and cell.value != None:
                    max_height = cell_height 
            except:
                pass
        sheet.row_dimensions[row[0].row].height = max_height

    locale.setlocale(locale.LC_TIME, 'en_US')

    start_of_week = date - timedelta(days=date.weekday())

    end_of_week = start_of_week + timedelta(days=6)

    start_date_str = start_of_week.strftime("%d")
    end_date_str = end_of_week.strftime("%d") + " " + start_of_week.strftime("%B")
    
    month = end_date_str.split(" ")[1]
    if month in month_fr:
        end_date_str = end_date_str.replace(month, month_fr[month]).replace(" ", "_")
    
    d_rot = f"du_{start_date_str}_au_{end_date_str}_{str(date.year)}"

    locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')

    classeur.save(os.path.join(path, "rotation_"+d_rot+".xlsx"))




#gérer formattage de la feuille
#write_planning(get_planning(absolute_path, datetime.now() + timedelta(days=-7), buses_list, spe_drivers), os.path.dirname(script_path))

